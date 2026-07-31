from django.core.files.uploadedfile import UploadedFile
from django.shortcuts import render
from django.http import HttpResponse
from django.core.handlers.wsgi import WSGIRequest
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from dashboard_meet_je_stad.form.mailchimp_form import MailchimpForm
import os
import sys
from django.apps import apps
import openpyxl
import re
from dashboard_meet_je_stad.repository.sensor_repository import SensorRepository
from dashboard_meet_je_stad.repository.user_repository import UserRepository


class MailchimpView:

    def __init__(self):
        path = os.path.dirname(apps.get_app_config('dashboard_meet_je_stad').path)
        if sys.argv[1:2] == ['test']:
            self.path_data = path + '/tests/data/'
        else:
            self.path_data = path + '/data/'
        self.sensor_repository = SensorRepository()
        self.user_repository = UserRepository()

    @method_decorator(staff_member_required)
    def index(self, request: WSGIRequest) -> HttpResponse:

        form  = MailchimpForm()
        message = ''
        if request.method == "POST":
            form = MailchimpForm(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES["file"]
                if not isinstance(file, list) and self._validate(file, form):
                    with open(self.path_data + 'Stations tbv Mailchimp.xlsx', "wb+") as destination:
                        for chunk in file.chunks():
                            destination.write(chunk)
                    destination.close()
                    self._process(file)
                    message = 'Inlezen is gelukt.'

        return render(request, 'admin/mailchimp.html', {'form': form, 'message': message})

    def _validate(self, file: UploadedFile, form: MailchimpForm) -> bool:
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active
        message = 'Het bestand is ongeldig.'
        if sheet_obj is None:
            form.add_error('file', message)
            return False
        for i in range(1, sheet_obj.max_row + 1):
            row = []
            for j in range(1, sheet_obj.max_column + 1):
                row.append(sheet_obj.cell(row=i, column=j).value)
            if len(row) != 3:
                form.add_error('file', message)
                return False
            regex = r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,7}"
            if isinstance(row[0], int) and row[1] is not None and re.fullmatch(regex, row[1]):
                return True
        form.add_error('file', message)

        return False

    def _process(self, file: UploadedFile):
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active
        if sheet_obj is None:
            return
        for i in range(1, sheet_obj.max_row + 1):
            row = []
            for j in range(1, sheet_obj.max_column + 1):
                row.append(sheet_obj.cell(row=i, column=j).value)
            regex = r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,7}"
            if isinstance(row[0], int) and row[1] is not None and re.fullmatch(regex, row[1]):
                user = self.user_repository.find_by_email(row[1])
                sensors = self.sensor_repository.find_all()
                if user is not None and row[0] in sensors:
                    self.user_repository.save_dashboard_user(user, row[0])
