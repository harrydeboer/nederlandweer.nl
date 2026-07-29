from django.contrib.auth.models import User
from typing import List
import openpyxl
from dashboard_meet_je_stad.models import DashboardUser
import os
import sys
from django.apps import apps


class UserRepository:

    def __init__(self):
        path = os.path.dirname(apps.get_app_config('dashboard_meet_je_stad').path)
        if sys.argv[1:2] == ['test']:
            self.path_data = path + '/tests/'
        else:
            self.path_data = path + '/data/'

    def get(self, user_id: int) -> User:

        return User.objects.get(pk=user_id)

    def find_all(self) -> List[User]:
        return list(User.objects.all())

    def find_by_username(self, username: str) -> User | None:
        return User.objects.filter(username=username).first()

    def find_by_email(self, email: str) -> User | None:
        return User.objects.filter(email=email).first()

    def create(self, user: User, password: str) -> User:
        user.set_password(password)
        user.save()
        wb_obj = openpyxl.load_workbook(self.path_data + 'Stations tbv Mailchimp.xlsx')
        sheet_obj = wb_obj.active
        rows = []
        assign = False
        sensor_id = None
        if sheet_obj is None:
            return user
        for i in range(1, sheet_obj.max_row + 1):
            row = []
            for j in range(1, sheet_obj.max_column + 1):
                row.append(sheet_obj.cell(row=i, column=j).value)
            if row[1] == user.email and isinstance(row[0], int):
                sensor_id = int(row[0])
                assign = True
            rows.append(row)
        if assign and sensor_id is not None:
            dashboard_user = DashboardUser()
            dashboard_user.set_sensor_id(sensor_id)
            dashboard_user.set_user(user)
            dashboard_user.save()

        return user

    def update(self, user: User, password: str | None = None) -> User:
        if password is not None:
            user.set_password(password)
        user.save()
        return user

    def delete(self, user: User) -> None:
        user.delete()

    def save_dashboard_user(self, user: User, sensor_id: int):
        dashboard_user = DashboardUser.objects.filter(_user=user, _sensor_id=sensor_id).first()
        if dashboard_user is not None:
            dashboard_user.set_sensor_id(sensor_id)
            dashboard_user.save()
        else:
            dashboard_user = DashboardUser()
            dashboard_user.set_sensor_id(sensor_id)
            dashboard_user.set_user(user)
            dashboard_user.save()
